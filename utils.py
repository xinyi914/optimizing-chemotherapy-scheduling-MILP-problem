import gurobipy as gp
from gurobipy import GRB
import scipy.io
import numpy as np
import math
from collections import defaultdict
from openpyxl import Workbook, load_workbook
import subprocess
import os
import time
import matplotlib.pyplot as plt

from parameter_sets import gb_par_sets,fs_par_sets,sc_par_sets
from necessary_data import u,v,days,data

# get the parameter set
def get_par_set(gb_setNo,fs_setNo,sc_setNo):
    return gb_par_sets[gb_setNo-1],fs_par_sets[fs_setNo-1],sc_par_sets[sc_setNo-1]

####################################################################
#
# Functions for training
#
####################################################################
# assign all appointments to each bed
# here var is something like different parameter sets
def first_step(instances,appo,v,u,mu,cov,alpha,weight_delta,weight_f, weight_delta_kb,gb_par_set,fs_par_set,patient_type = [30,60,120,180,240,300,360],B=14,K=7):
#     instances = "day1"
#     appo = [15,9,8,5,8,2,0]
#     print(appo)

    a= []
    
    for i in range(len(patient_type)):
        a.extend([patient_type[i]]*appo[i])
#     print(a_ordered)
#     a = np.random.permutation(a_ordered)
#     print(a)

    gamma = np.zeros([K,B])
    for b in range(B):
        for k in range(K):
            gamma[k,b] = np.sum(v[b]-u[b] == patient_type[k])

    phi = np.array(appo)

    omega = np.zeros([B,1])
    for b in range(B):
        omega[b] = np.sum(v[b]-u[b])

    model = gp.Model("model3_firstStep")
    model.setParam("TimeLimit",5000)

    model.setParam("LogFile",f"model3/firstStep/log/{instances}_alpha{alpha}_wd{weight_delta}_wf{weight_f}_wdk{weight_delta_kb}_set{gb_par_set}-{fs_par_set}.log")
    f = model.addMVar((K,B),vtype = GRB.INTEGER, name="f") # the number of appointment of type k to bed b
    beta = model.addMVar((K,B),vtype = GRB.BINARY, name="beta") # the type k in bed b is assigned
    theta = model.addMVar(B,vtype = GRB.BINARY, name="theta") # the bed b is assigned
    delta = model.addMVar((K,B),vtype = GRB.INTEGER, name="delta")
    # sqrt_term = model.addMVar(B,vtype=GRB.CONTINUOUS, name = "squareroot")

    td = 0

    for k in range(K):
    #     td = td + (1 + 6/k) * delta[k,:].sum()
        td = td + math.pow(weight_delta,k+1) * delta[k,:].sum()


    model.addConstr(delta >= f - gamma * beta)
    model.addConstr(delta >= gamma * beta - f)
    model.addConstr(f >= 0)  
    model.addConstr(theta >= beta) 
    model.addConstr(beta <= f) 


    # for k in range(K):
    #     for b in range(B):
    #         model.addConstr(theta[b] >= beta[k,b])


    for b in range(B):
        lhs = model.addVar(name=f"lhs_{b}")
        model.addConstr(lhs == ((1-alpha)/alpha) * f[:,b] @ cov @ f[:,b])

        rhs = model.addVar(name=f"rhs_{b}")
        model.addConstr(rhs == omega[b] * theta[b] - mu @ f[:,b] )

        model.addConstr(lhs <= rhs*rhs)

    for k in range(K):
        model.addConstr(f[k,:].sum() <= phi[k])


    model.setObjective(weight_delta_kb * td - weight_f * f.sum() + theta.sum() - beta.sum(), GRB.MINIMIZE)

    model.optimize()
    
    np.savez(f"model3/firstStep/result/{instances}_alpha{alpha}_wd{weight_delta}_wf{weight_f}_wdk{weight_delta_kb}_set{gb_par_set}-{fs_par_set}.npz",
             f = f.X, beta = beta.X, theta = theta.X,delta = delta.X, obj_value = np.array(model.Objval))
    
    print(f"file model3/firstStep/result/{instances}_alpha{alpha}_wd{weight_delta}_wf{weight_f}_wdk{weight_delta_kb}_set{gb_par_set}-{fs_par_set}.npz saved")
    
    return f.X,beta.X,theta.X,delta.X

def first_step_determined(instances,appo,v,u,mu,cov,alpha,weight_delta,weight_f, weight_delta_kb,gb_par_set,fs_par_set,patient_type = [30,60,120,180,240,300,360],B=14,K=7):
#     instances = "day1"
#     appo = [15,9,8,5,8,2,0]
#     print(appo)

    a= []
    
    for i in range(len(patient_type)):
        a.extend([patient_type[i]]*appo[i])
#     print(a_ordered)
#     a = np.random.permutation(a_ordered)
#     print(a)

    gamma = np.zeros([K,B])
    for b in range(B):
        for k in range(K):
            gamma[k,b] = np.sum(v[b]-u[b] == patient_type[k])

    phi = np.array(appo)

    omega = np.zeros([B,1])
    for b in range(B):
        omega[b] = np.sum(v[b]-u[b])

    model = gp.Model("model3_firstStep")
    model.setParam("TimeLimit",5000)
    model.setParam("LogFile",f"model3/firstStep/log/{instances}_alpha{alpha}_wd{weight_delta}_wf{weight_f}_wdk{weight_delta_kb}_set{gb_par_set}-{fs_par_set}.log")
   
    f = model.addMVar((K,B),vtype = GRB.INTEGER, name="f") # the number of appointment of type k to bed b
    beta = model.addMVar((K,B),vtype = GRB.BINARY, name="beta") # the type k in bed b is assigned
    theta = model.addMVar(B,vtype = GRB.BINARY, name="theta") # the bed b is assigned
    delta = model.addMVar((K,B),vtype = GRB.INTEGER, name="delta")
    # sqrt_term = model.addMVar(B,vtype=GRB.CONTINUOUS, name = "squareroot")

    td = 0

    for k in range(K):
    #     td = td + (1 + 6/k) * delta[k,:].sum()
        td = td + math.pow(weight_delta,k+1) * delta[k,:].sum()


    model.addConstr(delta >= f - gamma * beta)
    model.addConstr(delta >= gamma * beta - f)
    model.addConstr(f >= 0)  
    model.addConstr(theta >= beta) 
    model.addConstr(beta <= f) 


    # for k in range(K):
    #     for b in range(B):
    #         model.addConstr(theta[b] >= beta[k,b])


    for b in range(B):
        model.addConstr(mu @ f[:,b] <= omega[b]*theta[b])

    for k in range(K):
        model.addConstr(f[k,:].sum() <= phi[k])


    model.setObjective(weight_delta_kb * td - weight_f * f.sum() + theta.sum() - beta.sum(), GRB.MINIMIZE)

    model.optimize()
    
    np.savez(f"model3/firstStep/result/{instances}_alpha{alpha}_wd{weight_delta}_wf{weight_f}_wdk{weight_delta_kb}_set{gb_par_set}-{fs_par_set}.npz",
             f = f.X, beta = beta.X, theta = theta.X,delta = delta.X, obj_value = np.array(model.Objval))
    
    print(f"file model3/firstStep/result/{instances}_alpha{alpha}_wd{weight_delta}_wf{weight_f}_wdk{weight_delta_kb}_set{gb_par_set}-{fs_par_set}.npz saved")
    
    return f.X,beta.X,theta.X,delta.X
    

# return the objective value related to override policy
def get_override_objective(y,z,p,r,u,v,n,m):
    
    ob_policy_3 = np.zeros([n,m])
    for i in range(n):
        for j in range(m):
            ob_policy_3[i,j] = p[i,j] / (v[j]-u[j])
    ob_override = np.sum(y) + np.sum(z) - np.sum(ob_policy_3)-100*np.sum(r)
    
    return ob_override

# return the objective value related to override policy
def get_waitIdleTime_objective(lamda,beta,mui,gamma,second_mom,n):
    
    expr2 = np.zeros((n))
    for i in range(n):
        expr2[i] = beta[i]*mui[i]+gamma[i]*second_mom[i]
    ob_waitIdleTime = np.sum(lamda) + np.sum(expr2)
    
    return ob_waitIdleTime

# for 1 instance 1 bed
def second_step(instances,b,app,mb,T,u,v,M,mu,variances,
                wc,ic,wc_out,low_bound,up_bound,gb_par_set,fs_par_set,sc_par_set,y_weight,z_weight,x_weight,
                patient_type = [30,60,120,180,240,300,360],epsilon = 0.05,nu_ub = 1000000,nu_lb = -1000000,e_weight = 1,determined=False):
    
    print("nu_ub",nu_ub)
    print("nu_lb",nu_lb)
    a = []
    for i in range(len(patient_type)):
        a.extend([patient_type[i]]*int(app[i]))
    a = np.array(a)
  
    m = mb
    n = int(np.sum(app))
    pi = np.zeros((n+1,n+1))
    wait_cost = np.ones(n+1)*wc
    wait_cost[n] = wc_out
    idle_cost = np.ones(n)*ic

    pi = np.zeros((n+1,n+1))
    for i in range(n):
        for j in range(i,n):
            tau_sum = np.sum(wait_cost[i+1:j+1])
            pi[i,j] = -idle_cost[j]+ tau_sum

        pi[i,n] = wait_cost[n]+tau_sum
    
    patient_type_lowb = []
    patient_type_upb = []
    for i in patient_type:
        patient_type_lowb.append(i/low_bound)
        patient_type_upb.append(up_bound*i)
    #     patient_type_lowb.append(i)
    #     patient_type_upb.append(i)
    
 
    a_l = []
    a_u = []
    for i in a:
        idx = np.where(patient_type == i)[0][0]
        a_l.extend([patient_type_lowb[idx]])
        a_u.extend([patient_type_upb[idx]])
    a_l = np.array(a_l)
    a_u = np.array(a_u)
    
#     nu_ub = 1000000
#     nu_lb = -1000000
    nu_u = np.ones((n,n+1,n))*nu_ub
    nu_l = np.ones((n,n+1,n))*nu_lb
    
    mui = []
    second_mom = []
    for i in a:
        idx = np.where(patient_type == i)[0][0]
        mui.extend([mu[idx]])
        second_mom.extend([variances[idx]+mu[idx]**2])
    mui = np.array(mui)
    second_mom = np.array(second_mom)
    
    low_b = np.min(a)
    
#     e_weight = 1
    
    model = gp.Model(f"model3_secondStep_{b}")
    model.setParam("TimeLimit",500)
    # model.setParam("DualReductions",0)
    if not determined:
        model.setParam("LogFile",f"model3/secondStep/log/{instances}b{b}_wc{wc}_ic{ic}_wco{wc_out}_lb{low_bound}_ub{up_bound}_ewei{e_weight}_set{gb_par_set}-{fs_par_set}-{sc_par_set}_updated.log")
    else:
        model.setParam("LogFile",f"model3/secondStep/log/{instances}b{b}_wc{wc}_ic{ic}_wco{wc_out}_lb{low_bound}_ub{up_bound}_ewei{e_weight}_set{gb_par_set}-{fs_par_set}-{sc_par_set}_updated_determined.log")

    w = model.addMVar((n,n),vtype = GRB.BINARY,name = "w")
    z = model.addMVar((n,m),vtype = GRB.BINARY,name = "z")
    y = model.addMVar((m,n),vtype = GRB.BINARY,name = "y")
    x = model.addMVar((n,m),vtype = GRB.BINARY,name = "x")
    r = model.addMVar((n),vtype = GRB.BINARY,name = "r")
    g = model.addMVar((n),vtype = GRB.BINARY,name = "g")
    h = model.addMVar((m),vtype = GRB.BINARY,name = "h")
    q = model.addMVar((m,n),vtype = GRB.BINARY,name = "q")
    s = model.addMVar((n),vtype = GRB.CONTINUOUS, lb = 0,ub = T, name = "s")
    t = model.addMVar((n),vtype = GRB.CONTINUOUS,lb = 0,ub = T, name = "t")
    p = model.addMVar((n,m),vtype = GRB.CONTINUOUS, name = "p")
    d = model.addMVar((n),vtype = GRB.CONTINUOUS,lb = 0,ub = T, name = "d")
    lamda = model.addMVar((n+1),lb = -GRB.INFINITY,ub=GRB.INFINITY,vtype = GRB.CONTINUOUS, name = "lambda")
    beta = model.addMVar((n),lb = -GRB.INFINITY,ub=GRB.INFINITY,vtype = GRB.CONTINUOUS, name = "beta")
    kappa = model.addMVar((n,n+1,n),lb = -GRB.INFINITY,ub=GRB.INFINITY,vtype = GRB.CONTINUOUS, name = "kappa")
    nu =  model.addMVar((n+1,n+1,n),lb = -GRB.INFINITY,ub=GRB.INFINITY,vtype = GRB.CONTINUOUS, name = "nu")
    zeta = model.addMVar((n,n+1,n),lb = -GRB.INFINITY,ub=GRB.INFINITY,vtype = GRB.CONTINUOUS, name = "zeta")
    phi = model.addMVar((n,n+1,n),lb = -GRB.INFINITY,ub=GRB.INFINITY,vtype = GRB.CONTINUOUS, name = "phi")
    gamma = model.addMVar((n),lb = -GRB.INFINITY,ub=GRB.INFINITY,vtype = GRB.CONTINUOUS, name = "gamma")
    delta_l = model.addMVar((n,n+1,n),lb = -GRB.INFINITY,ub=GRB.INFINITY,vtype = GRB.CONTINUOUS, name = "delta_low")
    delta_u = model.addMVar((n,n+1,n),lb = -GRB.INFINITY,ub=GRB.INFINITY,vtype = GRB.CONTINUOUS, name = "delta_up")
    

 
                
    model.addConstr(s[0] >= u[0])


    for i in range(1,n):
        model.addConstr(s[i] >= t[i-1]) 
        model.addConstr(t[i-1] == s[i-1] + d[i-1])

#     model.addConstr(t[n-1] <= T)
    model.addConstr(t[n-1] == s[n-1] + d[n-1])
        
    expr = gp.MLinExpr.zeros((n,m))
    for i in range(n):
        for j in range(m):
            expr[i,j] = p[i,j] / (v[j]-u[j])
    
    for i in range(n):
        model.addConstr(w[i,:].sum() == 1)
        model.addConstr(w[:,i].sum() == 1)
        

    for i in range(n):
        for j in range(m):
            model.addConstr(p[i,j] >= x[i,j] * low_b)
            model.addConstr(p[i,j] >= d[i] + x[i,j]*T-T)
            model.addConstr(p[i,j] <= d[i] + x[i,j]*low_b-low_b)
            model.addConstr(p[i,j] <= x[i,j]*T)

    # constraints for z
    for i in range(1,n):
        for j in range(m):
            model.addConstr(t[i-1] - epsilon >= u[j] - M*(1-z[i,j]))
            model.addConstr(s[i] + epsilon <= v[j] + M*(1-z[i,j]))
            model.addConstr(z[i,j] <= r[i])
            model.addConstr(z[i,j] <= r[i-1])
        model.addConstr(z[i,:].sum() <= 1)

    for j in range(m):
        model.addConstr(z[0,j] == 0)
    
    # constraints for y when y = 1
    for i in range(n):
        for j in range(1,m):
            model.addConstr(s[i] + epsilon <= v[j-1] + M*(1-y[j,i]))
            model.addConstr(t[i] - epsilon >= v[j-1] - M*(1-y[j,i]))
            model.addConstr(y[j,i] <= r[i])

    for j in range(1,m):
        model.addConstr(y[j,:].sum() <= 1)
    
        
    # constraints for y when y = 0
    for i in range(n):
        for j in range(1,m):
            model.addConstr(s[i] + M*q[j,i] + M*y[j,i] >= v[j-1])
            model.addConstr(s[i] + M*q[j,i] + M*y[j,i] >= u[j]) # modified 6/23
            model.addConstr(t[i] - M*(1-q[j,i])-M*y[j,i] <= v[j-1])
            
#     for i in range(n):
#         for j in range(1,m):
#             model.addGenConstrMax(r[j-1, i], [v[j-1] - u[j], -1], name=f"max_z_{j}_{i}")
#             model.addConstr(y[j, i] <= 1 + r[j-1, i], name=f"y_constraint_{j}_{i}")

    for i in range(n):
         for j in range(1,m):
            model.addConstr(v[j-1] - u[j] >= -M*(1-y[j,i]))

    for i in range(n):
        model.addConstr(y[0,i] == 0)
        model.addConstr(q[0,i] == 0)
        
    for i in range(n):
        for j in range(1,m):
            model.addConstr(q[j,i] >= y[j,i])
            
    # constraints for g
    for i in range(1,n-1):
        model.addConstr(g[i] <= z[i,:].sum()+z[i+1,:].sum()+y[1:,i].sum())
    
    if n >= 2:
        model.addConstr(g[0] <= z[1,:].sum() + y[1:,0].sum()) 

    model.addConstr(g[n-1] <= z[n-1,:].sum() + y[1:,n-1].sum()) # problem
    
    for i in range (1,n):
        model.addConstr(g[i] >= z[i,:].sum()) #problem
        
    for i in range(n):
        for j in range(1,m):
            model.addConstr(g[i] >= y[j,i])
        model.addConstr(g[i] >= 0)
        model.addConstr(g[i] <= 1)

    # constraints for h
    for j in range(1,m-1):
        model.addConstr(h[j] <= z[1:,j].sum() + y[j,:].sum() + y[j+1,:].sum())
        
    model.addConstr(h[0] <= z[1:,0].sum() + y[1,:].sum())     

    model.addConstr(h[m-1] <= z[1:,m-1].sum() + y[m-1,:].sum())

    for j in range(m):
        for i in range(1,n):
            model.addConstr(h[j] >= z[i,j])
        model.addConstr(h[j] >= 0)
        model.addConstr(h[j] <= 1)

    for j in range(1,m):
        model.addConstr(h[j] >= y[j,:].sum())
        
    # constraints for x
    for i in range(n):
        for j in range(m):
            model.addConstr(x[i,j] <= 1-h[j])
            model.addConstr(x[i,j] <= 1-g[i])
            model.addConstr(x[i,j] <= r[i])
#             model.addConstr(s[i] >= u[j] - M[b]*(1-x[i,j]))
#             model.addConstr(t[i] <= v[j] + M[b]*(1-x[i,j]))
        model.addConstr(x[i,:].sum() <= r[i])
        model.addConstr(x[i,:].sum() == r[i] - g[i])

    for j in range(m):
        model.addConstr(x[:,j].sum() <= 1)
        
    # newly added
    for i in range(n):
        for j in range(m):
            model.addConstr(t[i] - epsilon >= u[j] - M*(1-x[i,j]))
            model.addConstr(s[i] + epsilon <= v[j] + M*(1-x[i,j]))

    # 4.10b
    expr47 = gp.MLinExpr.zeros((n+1,n+1))
    for i in range(n):
        for j in range(n+1):
            expr47[i,j] = pi[i,j]*(-d[i]+kappa[i,j,:].sum())
            
    for k in range(n+1):
        for j in range(k,n+1):   
            model.addConstr(lamda[k:j+1].sum() >= expr47[k:j+1,j].sum())
#             model.addConstr(lamda[k:j].sum() >= expr47[k:j,j].sum())
            
    # 4.10c
    for i in range(n):
        for j in range(i,n+1):
            if pi[i,j] > 0:
                for tau in range(n):
                    model.addConstr(nu[i,j,tau] >= -delta_l[i,j,tau]*a_l[tau]+delta_u[i,j,tau]*a_u[tau]+zeta[i,j,tau])
            elif pi[i,j] < 0:
                for tau in range(n):
                    model.addConstr(nu[i,j,tau] <= delta_l[i,j,tau]*a_l[tau]-delta_u[i,j,tau]*a_u[tau]-zeta[i,j,tau])
            
    # 4.6b,4.8b
    for i in range(n):
        for j in range(i,n+1):
            if pi[i,j] > 0: 
                for tau in range(n):
                    rhs0 = model.addVar()
                    rhs1 = model.addVar()
                    lhs = model.addVar()
                    model.addConstr(rhs0 == delta_l[i,j,tau] - delta_u[i,j,tau] + 1 - beta[tau] / pi[i,j])
                    model.addConstr(rhs1 == phi[i,j,tau] - zeta[i,j,tau])
                    model.addConstr(lhs == phi[i,j,tau] + zeta[i,j,tau])
                    
                    quad_expr = rhs0 * rhs0 + rhs1 * rhs1
                    model.addQConstr(quad_expr <= lhs * lhs, name=f"soc_{i}_{j}_{tau}")
            elif pi[i,j] < 0:
                for tau in range(n):
                    rhs0 = model.addVar()
                    rhs1 = model.addVar()
                    lhs = model.addVar()
            
                    model.addConstr(rhs0 == -delta_l[i,j,tau]+delta_u[i,j,tau]+1-beta[tau]/pi[i,j])
                    model.addConstr(rhs1 == phi[i,j,tau] - zeta[i,j,tau])
                    model.addConstr(lhs == phi[i,j,tau] + zeta[i,j,tau])
                    
                    quad_expr = rhs0 * rhs0 + rhs1 * rhs1
                    model.addQConstr(quad_expr <= lhs * lhs, name=f"soc_{i}_{j}_{tau}")
                
    # 4.5b,4.7b
    for i in range(n):
        for j in range(i,n+1):
            if pi[i,j] > 0:
                for tau in range(n):
                    model.addConstr(phi[i,j,tau] <= gamma[tau]/pi[i,j])
            elif pi[i,j] < 0:
                for tau in range(n):
                    model.addConstr(phi[i,j,tau] <= -gamma[tau]/pi[i,j])
    # 4.9a,b
    for i in range(n):
        for j in range(n+1):
            for tau in range(n):
                model.addConstr(kappa[i,j,tau] >= nu[i,j,tau] + (w[i,tau]-1)*nu_u[i,j,tau])
                model.addConstr(kappa[i,j,tau] >= nu_l[i,j,tau]*w[i,tau])
                model.addConstr(kappa[i,j,tau] <= nu_u[i,j,tau]*w[i,tau])
                model.addConstr(kappa[i,j,tau] <= nu[i,j,tau]+(w[i,tau]-1)*nu_l[i,j,tau])
                
    # 2.1c
    for i in range(n):
        model.addConstr(gamma[i] >= 0)
    #4.3d
    for tau in range(n):
        model.addConstr(nu[n,n,tau] == 0)
    #4.4c
    for i in range(n):
        for j in range(i,n+1):
            for tau in range(n):
                model.addConstr(phi[i,j,tau] >= 0)
                model.addConstr(delta_l[i,j,tau] >= 0)
                model.addConstr(delta_u[i,j,tau] >= 0)
    expr2 = gp.MLinExpr.zeros((n))
    for i in range(n):
        expr2[i] = beta[i]*mui[i]+gamma[i]*second_mom[i]
    model.setObjective(y_weight*y.sum() + z_weight*z.sum()-x_weight*expr.sum()-100*r.sum()+e_weight*(lamda.sum()+expr2.sum()),GRB.MINIMIZE)
#     model.setObjective(y.sum() + z.sum()-expr.sum()+e_weight*(lamda.sum()+expr2.sum()),GRB.MINIMIZE)
    model.optimize()
    
    assin = w.X@a
    
#     print(y.X)
#     print(z.X)
#     print(p.X)
    obj_override = get_override_objective(y.X,z.X,p.X,r.X, u,v,n,m)
    obj_waitIldeTime = get_waitIdleTime_objective(lamda.X,beta.X,mui,gamma.X,second_mom,n)

    if not determined:
        np.savez(f"model3/secondStep/result/{instances}b{b}_wc{wc}_ic{ic}_wco{wc_out}_lb{low_bound}_ub{up_bound}_ewei{e_weight}_set{gb_par_set}-{fs_par_set}-{sc_par_set}_updated.npz",
             d=d.X,w=w.X,z=z.X,y=y.X,x=x.X,g=g.X,h=h.X,q=q.X,s=s.X,t=t.X,p=p.X,r=r.X,lamda=lamda.X,beta=beta.X,kappa=kappa.X,
            nu = nu.X,zeta=zeta.X,phi=phi.X,gamma=gamma.X,delta_l=delta_l.X,delta_u=delta_u.X,app = assin, 
             obj_value = np.array(model.Objval),obj_override = np.array(obj_override),obj_waitIldeTime = np.array(obj_waitIldeTime))
    
        print(f"result saved in model3/secondStep/result/{instances}b{b}_wc{wc}_ic{ic}_wco{wc_out}_lb{low_bound}_ub{up_bound}_ewei{e_weight}_set{gb_par_set}-{fs_par_set}-{sc_par_set}_updated.npz")

    else:
        np.savez(f"model3/secondStep/result/{instances}b{b}_wc{wc}_ic{ic}_wco{wc_out}_lb{low_bound}_ub{up_bound}_ewei{e_weight}_set{gb_par_set}-{fs_par_set}-{sc_par_set}_updated_determined.npz",
             d=d.X,w=w.X,z=z.X,y=y.X,x=x.X,g=g.X,h=h.X,q=q.X,s=s.X,t=t.X,p=p.X,r=r.X,lamda=lamda.X,beta=beta.X,kappa=kappa.X,
            nu = nu.X,zeta=zeta.X,phi=phi.X,gamma=gamma.X,delta_l=delta_l.X,delta_u=delta_u.X,app = assin, 
             obj_value = np.array(model.Objval),obj_override = np.array(obj_override),obj_waitIldeTime = np.array(obj_waitIldeTime))
    
        print(f"result saved in model3/secondStep/result/{instances}b{b}_wc{wc}_ic{ic}_wco{wc_out}_lb{low_bound}_ub{up_bound}_ewei{e_weight}_set{gb_par_set}-{fs_par_set}-{sc_par_set}_updated_determined.npz")


####################################################################
#
# Functions for testing
#
####################################################################

# generate out of sample data
def generate_out_sample_data(mean,var,size = 1,rand_seed = None):
    if rand_seed is not None:
        np.random.seed(rand_seed)
    sigma2 = np.log(1 + var / mean**2)
    mu = np.log(mean) - 0.5 * sigma2
    sigma = np.sqrt(sigma2)

    return np.random.lognormal(mean = mu,sigma = sigma,size = size)

# load second step result
def second_load_result(instances,b,wc,ic,wc_out,lb,ub,e_weight,gb_par_set,fs_par_set,sc_par_set,updated=True,determined=False):
    if updated:
        if not determined:
            path = f"model3/secondStep/result/{instances}b{b}_wc{wc}_ic{ic}_wco{wc_out}_lb{lb}_ub{ub}_ewei{e_weight}_set{gb_par_set}-{fs_par_set}-{sc_par_set}_updated.npz"
        else:
            path = f"model3/secondStep/result/{instances}b{b}_wc{wc}_ic{ic}_wco{wc_out}_lb{lb}_ub{ub}_ewei{e_weight}_set{gb_par_set}-{fs_par_set}-{sc_par_set}_updated_determined.npz"
            print()

    else:
        path = f"model3/secondStep/result/{instances}b{b}_wc{wc}_ic{ic}_wco{wc_out}_lb{lb}_ub{ub}_ewei{e_weight}_set{gb_par_set}-{fs_par_set}-{sc_par_set}.npz"
    data = np.load(path)
    return data

# get second step statistics
def get_second_step_stat(B,mb,gc_no,fs_no,sc_no,simulate_times,rand_seed,out,sample_data,determined=False):
    gb_par_set,fs_par_set,sc_par_set = get_par_set(gc_no,fs_no,sc_no)
    updated = True
    ins = ["day1","day2","day3","day4","day5","day6","day7","day8","day9","day10","day11","day12","day13","day14","day15","day16","day17","day18","day19","day20","day21","day22"]
    wait_times = {}
    idle_times = {}
    over_times = {}
    override_policy1s = {}
    override_policy2s = {}
    override_policy3s = {}
    patient_type = np.array(gb_par_set["patient_type"])
    # simulate_times = 1000
    # rand_seed = 42
    np.random.seed(rand_seed)
#     for i in [3,11]:
    for i in range(len(ins)):
        print("---------------------------------------------------------",ins[i])
        o1_count = 0
        o2_count = 0
        o3_count = 0
        instances = ins[i]
        wait_times_day = []
        idle_times_day = []
        bed_overtimes_day = []
        for run_time in range(simulate_times):
#             print("run_time ", run_time)
            wait_time = 0
            idle_time = 0
            bed_over_time = 0

            for b in range(B):
    #             print("bed ",b)
                if not determined:
                    data = second_load_result(instances,b,sc_par_set['wc'],sc_par_set['ic'],sc_par_set['wc_out'],sc_par_set['lb'],sc_par_set['ub'],sc_par_set['e_weight'],gb_par_set['gb_set_no'],fs_par_set['fs_set_no'],sc_par_set['sc_set_no'],updated)
                else:
                    data = second_load_result(instances,b,sc_par_set['wc'],sc_par_set['ic'],sc_par_set['wc_out'],sc_par_set['lb'],sc_par_set['ub'],sc_par_set['e_weight'],gb_par_set['gb_set_no'],fs_par_set['fs_set_no'],sc_par_set['sc_set_no'],updated,determined)
                assign = data["d"]
                appo = data["app"]
                s = data["s"]
                t = data["t"]
                simu_data = []
#                 print("assign ",assign)
#                 print("apppo ",appo)
                # randomly choose the sample from corresponding type
                for a in range(len(assign)):
#                     print("appo[a] ",appo[a])
#                     print("appo[a] ",int(appo[a]))
#                     print("patient_type ",patient_type)
                   
                    sample_idx = np.where(np.isclose(patient_type, appo[a], atol=1e-5))[0][0]
                    test_data = sample_data[sample_idx]
                    # if out:
                    #     test_data = out_sample_data[sample_idx]
                    # else:
                    #     test_data = in_sample_data[sample_idx]
                    random_choose_data = np.random.choice(test_data.flatten())
#                     while random_choose_data < appo[a]/3 or random_choose_data > appo[a]*4:
# #                        print("rechoose: ",random_choose_data)
#                         random_choose_data = np.random.choice(test_data)
                    simu_data.append(random_choose_data)
                
    #             print(simu_data)

                # simulate real start and end time of each assignment
                real_start = []
                real_end = []
                for idx in range(len(simu_data)):
                    if idx == 0:
                        real_start.append(s[idx])
                        real_end.append(s[idx] + simu_data[idx])
                    else:
                        if real_end[idx-1] > s[idx]:
                            real_start.append(real_end[idx-1])
                            real_end.append(real_end[idx-1] + simu_data[idx])
                        else:
                            real_start.append(s[idx])
                            real_end.append(s[idx] + simu_data[idx])

                # calculate wait_time,idle_time,over_time    
                for idx in range(len(simu_data)):
                    # calculate idle time
                    if real_end[idx] < t[idx]:
                        idle_time += (t[idx] - real_end[idx])

    #                     print("idle_time ",idle_time)

                    # calculate wait time
                    if real_start[idx] > s[idx]:
                        wait_time += (real_start[idx] - s[idx])
    #                     print("wait_time ",wait_time)
                    # calculate bed over time:
                    if idx == len(simu_data) - 1:
                        if real_end[idx] > t[idx]:
                            bed_over_time += (real_end[idx] - t[idx])
    #                         print("bed_over_time ",bed_over_time)

    #                 if t == 0:


    #                     if assign[t] > simu_data[t]: 
    #                         idle_time += (assign[t]-simu_data[t])

    #                 else:
    #                     if t < len(simu_data) - 1 or t > 0:
    #                         real_end = s[t-1] + simu_data[t-1]
    #                         if real_end > s[t]:
    #                             wait_time += (simu_data[t]-assign[t])
    #                     elif t == len(simu_data) - 1:
    #                         bed_over_time += simu_data[t]-assign[t]
    #             if run_time < 10:
    #                 print("appo ",appo)
    #                 print("assign ",assign)
    #                 print("simu_data ",simu_data)
    #                 print("s ",s)
    #                 print("t ",t)
    #                 print("real_start ",real_start)
    #                 print("real_end ",real_end)
    #                 print("idle_time ",idle_time)
    #                 print("wait_time ",wait_time)
    #                 print("bed_over_time ",bed_over_time)
            wait_times_day.append(wait_time)
            idle_times_day.append(idle_time)
            bed_overtimes_day.append(bed_over_time)


        for b in range(B):
            data = second_load_result(instances,b,sc_par_set['wc'],sc_par_set['ic'],sc_par_set['wc_out'],sc_par_set['lb'],sc_par_set['ub'],sc_par_set['e_weight'],gb_par_set['gb_set_no'],fs_par_set['fs_set_no'],sc_par_set['sc_set_no'],updated)
            o1 = data["x"]
            o2 = data["y"]
            o3 = data["z"]
            assign = data["d"]
            appo = data["app"]
            o1_count += np.sum(o1)
            o2_count += np.sum(o2)
            o3_count += np.sum(o3)
    #         print(o1_count)
    #         print(o2_count)
    #         print(o3_count)

            for a in range(len(assign)):
                for s in range(mb[b]):
                    if o1[a,s] > 0.1:
                        if assign[a]-0.5 <= appo[a] and assign[a]+0.5 >= appo[a]:
#                             print(instances," ",b)
#                             print("assign[a] ",assign[a])
#                             print("appo[a] ",appo[a])
#                             print("o1_count before: ",o1_count)
                            o1_count = o1_count - 1
#                             print("o1_count after: ",o1_count)

        wait_times[instances] = wait_times_day
        idle_times[instances] = idle_times_day
        over_times[instances] = bed_overtimes_day
        override_policy1s[instances] = o1_count
        override_policy2s[instances] = o2_count
        override_policy3s[instances] = o3_count
    if not determined:
        if out:
            np.savez(f"model3/secondStep/statistics/simTime{simulate_times}_seed{rand_seed}_set{gb_par_set['gb_set_no']}-{fs_par_set['fs_set_no']}-{sc_par_set['sc_set_no']}_updated.npz",
                    wait_times = wait_times,idle_times = idle_times,over_times=over_times,override_policy1s = override_policy1s,override_policy2s = override_policy2s,override_policy3s = override_policy3s)
            print(f"model3/secondStep/statistics/simTime{simulate_times}_seed{rand_seed}_set{gb_par_set['gb_set_no']}-{fs_par_set['fs_set_no']}-{sc_par_set['sc_set_no']}_updated.npz saved")

        else:
            np.savez(f"model3/secondStep/statistics/simTime{simulate_times}_seed{rand_seed}_set{gb_par_set['gb_set_no']}-{fs_par_set['fs_set_no']}-{sc_par_set['sc_set_no']}_updated_inSample.npz",
                    wait_times = wait_times,idle_times = idle_times,over_times=over_times,override_policy1s = override_policy1s,override_policy2s = override_policy2s,override_policy3s = override_policy3s)

            print(f"model3/secondStep/statistics/simTime{simulate_times}_seed{rand_seed}_set{gb_par_set['gb_set_no']}-{fs_par_set['fs_set_no']}-{sc_par_set['sc_set_no']}_updated_inSample.npz saved")
    else:
        if out:
            np.savez(f"model3/secondStep/statistics/simTime{simulate_times}_seed{rand_seed}_set{gb_par_set['gb_set_no']}-{fs_par_set['fs_set_no']}-{sc_par_set['sc_set_no']}_updated_determined.npz",
                    wait_times = wait_times,idle_times = idle_times,over_times=over_times,override_policy1s = override_policy1s,override_policy2s = override_policy2s,override_policy3s = override_policy3s)
            print(f"model3/secondStep/statistics/simTime{simulate_times}_seed{rand_seed}_set{gb_par_set['gb_set_no']}-{fs_par_set['fs_set_no']}-{sc_par_set['sc_set_no']}_updated_determined.npz saved")

        else:
            np.savez(f"model3/secondStep/statistics/simTime{simulate_times}_seed{rand_seed}_set{gb_par_set['gb_set_no']}-{fs_par_set['fs_set_no']}-{sc_par_set['sc_set_no']}_updated_inSample_determined.npz",
                    wait_times = wait_times,idle_times = idle_times,over_times=over_times,override_policy1s = override_policy1s,override_policy2s = override_policy2s,override_policy3s = override_policy3s)

            print(f"model3/secondStep/statistics/simTime{simulate_times}_seed{rand_seed}_set{gb_par_set['gb_set_no']}-{fs_par_set['fs_set_no']}-{sc_par_set['sc_set_no']}_updated_inSample_determined.npz saved")


####################################################################
#
# Functions for analysis and writing to excel
#
####################################################################

# functions for statistics of the first step
def get_fs_stat(gb_no,fs_no):
    perc_whole = []
    perc_30 = []
    perc_60 = []
    perc_120 = []
    perc_180 = []
    perc_240 = []
    perc_300 = []
    perc_360 = []
    app_counts = []
    ass_counts = []
    ins = ["day1","day2","day3","day4","day5","day6","day7","day8","day9","day10","day11","day12","day13","day14","day15","day16","day17","day18","day19","day20","day21","day22"]
    for i in range(len(ins)):
        instances = ins[i]
        gb_par_set,fs_par_set,_ = get_par_set(gb_no,fs_no,2)
        fs_path = f"model3/firstStep/result/{instances}_alpha{fs_par_set['alpha']}_wd{fs_par_set['weight_delta']}_wf{fs_par_set['weight_f']}_wdk{fs_par_set['weight_delta_kb']}_set{gb_par_set['gb_set_no']}-{fs_par_set['fs_set_no']}.npz"
        fs_result = np.load(fs_path)
        f_arr = fs_result["f"]
        ass_count = np.sum(f_arr)
        type_count = np.sum(f_arr,1)
        # print(type_count)
        app_list_count = sum(days[i])
        app_counts.append(app_list_count)
        ass_counts.append(ass_count)
        perc_whole.append(ass_count/app_list_count)
        perc_30.append(type_count[0]/days[i][0] if days[i][0] != 0 else None)
        perc_60.append(type_count[1]/days[i][1] if days[i][1] != 0 else None)
        perc_120.append(type_count[2]/days[i][2] if days[i][2] != 0 else None)
        perc_180.append(type_count[3]/days[i][3] if days[i][3] != 0 else None)
        perc_240.append(type_count[4]/days[i][4] if days[i][4] != 0 else None)
        perc_300.append(type_count[5]/days[i][5] if days[i][5] != 0 else None)
        perc_360.append(type_count[6]/days[i][6] if days[i][6] != 0 else None)
    return app_list_count, app_counts, perc_whole, perc_30, perc_60, perc_120, perc_180, perc_240,perc_300,perc_360
    
# visualize the assignment result in excel
def write_to_excel(data,b,instances):
    
    print("bed",b)
    subprocess.run(["osascript", "-e", 'tell application "Microsoft Excel" to quit saving yes'])
    
    result_col = ["L","O","R","U","X","AA","AD","AG","AJ","AM","AP","AS","AV","AY"]
    appo_col = ["M","P","S","V","Y","AB","AE","AH","AK","AN","AQ","AT","AW","AZ"]
    
    # write to excel
    current_path = os.getcwd()
    excel_file = f"{current_path}/model3_optimization_graph2.xlsm"
    wb = load_workbook(excel_file,keep_vba = True)
    ws = wb[instances]

    start_row = 2
    end_row = ws.max_row
    s = data["s"]
    t = data["t"]
    app = data["app"]
    
    # clear the column first
    for col in range(4,7):
        for row in range(start_row, end_row+1):
            ws.cell(row=row,column=col).value = None
    
    # write s
    for i,value in enumerate(s,start=start_row):
        ws.cell(row=i,column=4,value=int(value))


    # write t
    for i,value in enumerate(t,start=start_row):
        ws.cell(row=i,column=5,value=int(value))

    # write app
    for i,value in enumerate(app,start=start_row):
        ws.cell(row=i,column=6,value=int(value))

    wb.save(excel_file)
    wb.close()
    
    print("data and written and saved")
    time.sleep(2)
    
    result_macro_name = f"result_{result_col[b]}.ApplyConditionalFormatting"
    
    
    result_script = f'''
    tell application "Microsoft Excel"
        activate
        set wb to open workbook workbook file name (POSIX file "{excel_file}")
        tell worksheet "{instances}" of wb to activate
        run VB macro "{result_macro_name}"
        save workbook wb
        close workbook wb saving yes
    end tell
    '''
    
    subprocess.run(["osascript","-e",result_script])
    print("result macro applied")
    
    time.sleep(2)
    
    appo_macro_name = f"appo_{appo_col[b]}.ApplyConditionalFormatting"
    
    appo_script = f'''
    tell application "Microsoft Excel"
        activate
        set wb to open workbook workbook file name (POSIX file "{excel_file}")
        tell worksheet "{instances}" of wb to activate
        run VB macro "{appo_macro_name}"
        save workbook wb
        close workbook wb saving yes
    end tell
    '''
    subprocess.run(["osascript","-e",appo_script])
    print("appo mcaro applied")
    
    print("finished")

# get visual results
def get_visual_result(B,gc_no,fs_no,sc_no,instances):
    gb_par_set,fs_par_set,sc_par_set = get_par_set(gc_no,fs_no,sc_no)
    updated = True
#     ins = ["day1","day2","day3","day4","day5","day6","day7","day8","day9","day10","day11","day12","day13","day14","day15","day16","day17","day18","day19","day20","day21","day22"]
    for b in range(B):
    #             print("bed ",b)
        data = second_load_result(instances,b,sc_par_set['wc'],sc_par_set['ic'],sc_par_set['wc_out'],sc_par_set['lb'],sc_par_set['ub'],sc_par_set['e_weight'],gb_par_set['gb_set_no'],fs_par_set['fs_set_no'],sc_par_set['sc_set_no'],updated)
        write_to_excel(data,b,instances)

# load second step stat
def load_step2_stats(simulate_times,rand_seed,gb_par_set,fs_par_set,sc_par_set,out,determined=False):
    if not determined:
        if out:
            stat = np.load(f"model3/secondStep/statistics/simTime{simulate_times}_seed{rand_seed}_set{gb_par_set['gb_set_no']}-{fs_par_set['fs_set_no']}-{sc_par_set['sc_set_no']}_updated.npz",allow_pickle=True)
        else:
            stat = np.load(f"model3/secondStep/statistics/simTime{simulate_times}_seed{rand_seed}_set{gb_par_set['gb_set_no']}-{fs_par_set['fs_set_no']}-{sc_par_set['sc_set_no']}_updated_inSample.npz",allow_pickle=True)
    else:
        if out:
            stat = np.load(f"model3/secondStep/statistics/simTime{simulate_times}_seed{rand_seed}_set{gb_par_set['gb_set_no']}-{fs_par_set['fs_set_no']}-{sc_par_set['sc_set_no']}_updated_determined.npz",allow_pickle=True)
        else:
            stat = np.load(f"model3/secondStep/statistics/simTime{simulate_times}_seed{rand_seed}_set{gb_par_set['gb_set_no']}-{fs_par_set['fs_set_no']}-{sc_par_set['sc_set_no']}_updated_inSample_determined.npz",allow_pickle=True)

    wait_times = stat["wait_times"].item()
    idle_times = stat["idle_times"].item()
    over_times = stat["over_times"].item()
    override_policy1s = stat["override_policy1s"].item()
    override_policy2s = stat["override_policy2s"].item()
    override_policy3s = stat["override_policy3s"].item()
    return wait_times,idle_times,over_times,override_policy1s,override_policy2s,override_policy3s

# write stats into excel
def stats_to_excel(filename,wait_times,idle_times,over_times,override_policy1s,override_policy2s,override_policy3s,column_idx,start_row):

#     filename = "model3/secondStep/second_step_statistics.xlsx"
    wb = load_workbook(filename)
    ws = wb["Sheet1"]
    # write wait_time
#     column_idx = 2 #"B"

#     start_row = 4
    all_stat = [wait_times,idle_times,over_times,override_policy1s,override_policy2s,override_policy3s]
    # clear row first
    for col in range(column_idx, column_idx+6):  # columns D, E, F
        for row in range(start_row, start_row+22):
#         for row in [start_row+3,start_row+11]:
            print("col ",col)
            print("row ",row)
            ws.cell(row=row, column=col).value = None

    for j,d in enumerate(all_stat):
        for i,(key,value) in enumerate(d.items()):
            if j < 3:
                ws.cell(row = start_row + i,column = column_idx + j,value = np.mean(value))
            else:
                ws.cell(row = start_row + i,column = column_idx + j,value = value)
#     for j,d in enumerate(all_stat):
#             if j < 3:
#                 ws.cell(row = start_row + 3,column = column_idx + j,value = np.mean(d["day4"]))
#                 ws.cell(row = start_row + 11,column = column_idx + j,value = np.mean(d["day12"]))
#             else:
#                 ws.cell(row = start_row + 3,column = column_idx + j,value = d["day4"])
#                 ws.cell(row = start_row + 11,column = column_idx + j,value = d["day12"])

    wb.save(filename)
    wb.close()
    print("finished")
